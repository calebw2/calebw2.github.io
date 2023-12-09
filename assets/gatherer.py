from openpyxl import Workbook, load_workbook
import numpy as np 
import os
import pandas as pd


BASE_GREASES = ['8626 BASE GREASE ', '87 BASE GREASE (AX460)', 'AX 100 #2 BASE GREASE',
 				'AX100 #2 BASE GREASE', 'AX220 BASE GREASE', 'AX460 BASE GREASE', 
 				'CA 207MPC BASE', 'CLARETECH ECOCURVE TRACKSIDE BASE GREASE', 
 				'GEOTHERM 2020 BASE', 'HEAVY HAUL #1.5 BASE GREASE', 
 				'LI #2 STANDARD BASE (LI 220)', 'LI150 #2 STANDARD BASE', 'LITHIUM #3.5 BASE ', 
 				'LITHIUM COMPLEX STANDARD BASE', 'LOOM OIL BASE GREASE', 'LUCAS RED N TACKY BASE', 
 				'LUCAS WHITE LITHIUM BASE', 'RAIL KING HL 1.5 BASE GREASE', 'LX220', 'LX220 #2'
 				'LX150', 'SMG 500 #2 STANDARD BASE', 'SMG 700 #2 STANDARD BASE']
ALT_BASED_GR = ['ConAg Base Grease', 'Conag Base Concentrate', 'Lithium Complex Concentrate', 'RSC Base', 'Base Grease', 'SMG 500 Base', 'Base Grease (Pen 265)', 'Cup Calcium Grease #2 Base', 'R1 Base Grease (p60: 155)', 'Duraloc LD LIN26 Base', 'Enduron Base', 'Conag Base', 'B2 base Grease', 'Pilot Alpha #1.5 Base', 'SMG 700 Base Grease', 'RBL 110 #1.5 Base', 'SMG Base Conc', 'Lithium Base Concentrate', 'B2 Base Grease', 'Lith Base Conc', 'Lith Base Conce', 'R1 base grease, (pen 114)', 'Cotton picker base', 'Conc Lith Base Grease', 'Heavy Loom Oil Base Grease', 'CONCETRATE BASE', 'Thick base grease ', 'Li Complex Base Gr ', 'Thick Base Grease ', 'Mixed Base ', 'LI #2 Std Base', 'R2 Base Lithx Gr', 'Alum unknown base', 'LI #2 Base Conc', 'Rail King HL #1.5 Base Grease ', 'LI #2 Base Conc', '8626 Base Grease', '87 Base Grease', '87 Base Grease ', '87 Base Grease (AX460) ', 'AX100 Base Gr', 'AX100 Base Grease', 'AX220 #2 Base Grease', 'B-28 Base', 'b20 Base', 'B20 Base ', 'B20/B28 Base', 'B21 Base', 'b21 base ', 'B21 Base Gr', 'b21 base gr ', 'B28', 'B28 Base', 'B28 Base ', 'B28 base ', 'B28 BASE ', 'B28 Base Gr', 'b28 base gr', 'B28 Base Gr ', 'b28 base Gr ', 'B28 Base Grease', 'B28 Base Grease ', 'b28 baSE Grease ', 'B28 Bse Grease ', 'B28 from bin flush', 'BASE CONCENTRATE', 'Base Concentrate', 'Base Grease ', 'Base Grease Only', 'Base Pan', 'BK2 BASE', 'BK2 BASE  GR ', 'BK2 Base Gr', 'BK3 BASE', 'BK3 Base', 'Conag Concentrate', 'Conc Base', 'CPG Base Grease', 'Ecobase 1', 'Ecobase 2', 'FPL 121 Con Ag Base', 'FPL 121 Con Ag Base ', 'Geotherm 900 Base', 'Geotherm XG2 Base', 'Heavy Loom Oil Base Grease ', 'K3 Bse Gr', 'LI #2 Standard Base', 'LI #2 STD BASE', 'LI #2 Thick Base', 'LI 150 #2 Standard Base', 'LI 220 #2 Standard Base ', 'LI BASE #3.5', 'LI#2 Standard Base', 'LI#2 Std Base ', 'LI220 #2 Standard Base', 'LI220 #2 STD BASE', 'Lithium #3.5 Base', 'Lithium #3.5 Base Lot#: W10027I', 'Lithium 3.5 Base #3.5 V27205I', 'Lithium Base ', 'Lithium Base #3.5', 'lithium Base #3.5 2 Drums ', 'Lithium Base 3.5 ', 'Lithium Cmplex Std  B ', 'Lithium Complex Standard B', 'Lithium Complex Std', 'Lithium Complex Std B', 'Lithium Complex Std B k2', 'Lithium Complex Std Base', 'Lithium Complex STd Base ', 'Lithium Complex Std Base ', 'Lithium Complex Std. B', 'Lithium Standard Complex Std B', 'Loom Base', 'Loom Oil Base', 'Loom Oil Base Grease ', 'Lube a Boom #2 Base', 'Lucas Red N Tacky Base ', "Lucas Red N' Tacky Base", 'Lucas White Lith Base', 'Lucas White Lithium Base', 'Lucas White Lithium Base ', 'LX 150 #2 Standard Base', 'LX 220 #2 Standard Base', 'LX150 #2 Standard Base', 'lx150 std baSE Gr ', 'LX220 #2 Standard Base', 'PA #2 Base Grease', 'R1 Base', 'R1 Base ', 'R2 Base', 'R2 Base ', 'R2 Base Grease ', 'R2 Lithium Cmplex Base ', 'SCA 680 Base', 'SMG #2 Standard Base', 'SMG 500 # 2 Standard Base', 'SMG 500 Base Grease', 'SMG 500#2 Standard Base', 'SMG 700 #2 Base', 'Thermax 680 Base', 'Thick Base R2', 'White Lithium Base', 'XM 9590 150-2 Base ', 'XM 9590/150-2 Base']
SILICA = ['Aerosil R208', 'Aerosil 200', 'Dust Collector Silica ', 'Aerosil R 972 / Reolosil DM-1', 'Aerosil R 972', 'Aerosil ', 'Aerosil 972', 'Aerosil R927', 'Aerosil R972', 'Aerosil R972 / Reolosil DM 10', 'Aerosil R972 / Reolosil DM-1', 'Reolisil DM-10', 'Reolosil DM 10', 'Reolosil DM-10']
FLUSH = ['Dark Flush Oil', 'Tote #3 Flush Oil', 'Moly 3% EP GR LC 0 Flush', 'Flush Oil Tote #21', 'Flsuh', 'FPL 105 Flush Drums', '600N Flush', 'Red Trailer Flush w/R750', 'Moly 3\% EP GR LC 0 Flush', 'Flush oil (#7)', 'Termalene flush w/chevron', 'Termalene Flush w/chevron', 'FPL 123 Flush Oil ', 'Termalene Flush w/ Raffene', 'BK3 Flush', 'CA-207MPC FLUSH', 'Drum next to R1', 'Flush', 'Flush Oil', 'Flush oil', 'Flush Oil ', 'Flush oil ', 'flush oil ', 'Flush Tank', 'FPL 105 FLUSH', 'FPL 123 Con-Ag FLUSH ', 'FPL 123 COnAG Flush ', 'fpl 123 Flush Oil ', 'HD AC 180B FLUSH', 'HEAVY HAUL FLUSH', 'Heavy Haul Flush', 'Loom Base Flush', 'Loom Oil Flush', 'Loom Oil Flush ', 'Lubrisilk Flush', 'Lucas Marine Flush', 'LUCAS MARINE FLUSH', 'Oil Drained from Boiler', 'Raffene750 w/FPL 123 ', 'Red Flush Oil', 'Red Flush Tote', 'Silver 70 Flush', 'Silver Flush', 'SMG 500 with Char Flushed with R600', 'Trailer Red Flush Rework', 'White Flush', 'White Lith Flush', 'White Lithium Flush ']
THICC_BOIS = ['Focus 2500', 'SSI24', 'TomVis SSI 50', 'Focus PA 2500', 'SSI 50', 'Indopol H1900', 'Ssi 50', 'Petro Resin 25/ Kendex 0834', 'TPC 1350/Polybutene ', 'Polybutene 32', 'Lucant 1100', 'Lucant 110', 'FOCUS PA 2500', 'Daelimsynol 1100', 'Indopol H-100', 'Indopol H-100 w/ Silver 70 Rework', 'Indopol H-1900', 'Indopol H-300', 'Indopol with Silver 70', 'Polybutene 1300', 'Polybutene 2400', 'TPC 1350', 'Polymer Re work ', 'Polymer re work ', 'Polyurea Rework', 'Polyurea rework', 'Priolube 3986', 'Tom Vis SSI-24', 'TOmVis SSI 24', 'TomVis SSI 24', 'TOMVIS SSI 24 ', 'TomVis SSI-24', 'Tomvis SSI-24', 'TomVis SSI-24 ', 'TomVis SSI-50', 'TomVis SSI24', 'TomVis SSI50', 'TomViss SSi24 ', 'Petro Resin 25 / Kendex 0834', 'Petro Resin 25 Kendex 0834']
BASE_OILS = ['150 BRIGHT STOCK', 'CHEVRON 100', 'CHEVRON 220', 'CHEVRON 600', 'HYNAP 100', 'HYNAP 40', 
			 'KENDEX 0834', 'NESTE 3043', 'PAO 10', 'PAO 100', 'PAO 4', 'PAO 40', 'PAO 6', 'PAO 8', 
			 'RADIALUBE 7362', 'RAFFEENE 2000L', 'RAFFENE 1200L', 'RAFFENE 750L', 'SSI-24', 'SSI-50',
			 '150 BS', ]
ALT_BASED_OIL = ['Raffen 1200', 'Lab Rework Oil (additives)', 'Raffene 2500', 'PMX 200 Silicone 100 cST', 'DMS300-1000', 'Hyanp 100', 'SK120 ', 'RB PAO 40 ', 'Radialube ', 'SK 120', 'R750', '150 Brite Stock', 'Element 14 PDMS 1000', 'DMS300-350', 'RB PAO 8 w/ mineral oil 70', 'FPL- Tote 210', 'RB PAO 8', 'Mineral Oil 380', 'DMS300-100', '150 BS/ Renoil 2602', 'DMS300-10000', 'Mixed Oil from Lab', 'Hatcol 2941', 'Ucon LB 1715', 'UCON LB 1715', 'Ucon LB-1715','150 Bright Stock ', '150 Bright Stock Renol ', '150 Bright Stock/Renoil 260', '150 BS ', '150 BS Rework', '150 BS/Renoil 2602', 'Brite Stock 150', 'Chevron ', 'Chevron 100 / SK 120', 'Chevron 220 ', 'Chevron 600 ', 'Chevron 600 (holdout)', 'Chevron 600N', 'Chevron 7R', 'Exxon 150 Brite Stock', 'Exxon 150 BS', 'Holly 600 SN', 'Hynap ', 'Hynap 100 ', 'Hynap 2000', 'Radialube 7368', 'Radialube 7669', 'Raffene ', 'Raffene 1200', 'Raffene 1200L ', 'Raffene 2000L', 'raffene 750', 'Raffene 750', 'Raffene 750 L ', 'Raffene750L', 'Rafffene 750L ', 'RB PAO 100', 'RB PAO 4 ', 'RB PAO 40', 'RB PAO 6', 'RB PAO 8 ', 'Safety Kleen', 'Safety Kleen 120', 'Signal Fluids 100 ', 'Signal Fluids 100 N', 'Signal Fluids 100 Neutral', 'SJR Raffene ', 'SJR Raffene 750L', 'Spectra Syn Elite', 'Spectrasyn Elite 150', 'SpectraSyn Elite 65', 'Spectrasyn Elite 90', 'Spectrum V']

INSTRUCTIONS = ['Mix for 30  mins ', 'Grab 1 Gallon Can as Sample ', 'De Aerate 2Hrs', 'Dearate for 2 hrs ', 'Mill for 30 minutes', 'Get 1 Gallon Sample can ', 'Dearate for 2 HRS ', 'Set Mill hole to #7', 'Deaerate 1 Hour', 'Grab 1 Gallon Can Sample ', 'Mix & Recirculate for 45 min', 'De Aerate 2 Hrs ', 'Deaerate for 1 hours', 'Grab 1 Gal Can as Sample ', 'Resample ', 'Deaerate for 1 hr', 'De-Aerate 1hr', 'TRANSFER to BK2', 'Re Sample ', 'De Aerate for 45mins  ', 'Avoid blades', 'Change Mill setting to #2', 'DE AERATE  1 HR', 'DE AERATE 1 HR', 'DE AERATE 2 HRs', 'DE AERATE 2 HRS ', 'DE AERATE 2HRs', 'De aerate for 2 Hrs ', 'De Aerate for 2 Hrs ', 'De aerate for 2 HRS ', 'DE_AERATE 2 HRS', 'De-aerate 1 Hr', 'De-Aerate 1 hr', 'DE-AERATE 1 HR', 'DE-AERATE 1HR', 'De-Aerate 2 hr', 'DE-AERATE 2 HRs', 'DE-AERATE 2HRs', 'DE-AERATE 2HRS', 'DE-AERATE 3 HRs', 'De-aerate for 1 hour', 'De-aerate for 1 hr', 'DE-AERATE FOR 1 HR', 'De-aerate for 2 hours', 'De-aerate for 2 hrs', 'DE-AERATE FOR 2 HRS', 'DEAERATE 1 HR', 'Deaerate 1 Hr', 'DEAERATE 2 HRS', 'Deaerate 2 Hrs', 'DEAERATE 3 HRS', 'DEAERATE FOR 1 HR', 'Deaerate for 1 HR ', 'Deaerate for 1 Hr ', 'Deaerate for 2 hours', 'Deaerate for 2 Hr ', 'Deaerate for 2 Hrs ', 'Deaerate for 30 MINS  ', 'Dearate 2 Hrs', 'DEARATE FOR 1 HOUR', 'Dearate for 2 Hours', 'DEAREATE 2 HRS', 'Gal Needed', 'GET MILLED SAMPLE', 'Grab 1  Gallon can for final testing ', 'Grab 1 Gal Can as Sample', 'Grab 1 Gal can as sample ', 'Grab 1 Gal Can as sample ', 'Grab 1 Gal Can Sample ', 'Grab 1 gal sample', 'Grab 1 Gal Sample ', 'Grab 1 Gal Sample Can ', 'Grab 1 Gallon Can for Final testing ', 'Grab 1 Gallon Can for Final Testing ', 'Grab 1 Small Can as sample ', 'Grab 1-Gal Can for Final Testing', 'Grab 1-SMALL Can for Final Testing', 'Grab milled and unmilled sample', 'Grab Small Can as sample ', 'Main Assumption: "Gallons per 10 pts". It is assumed that you need 100 gallons per 10 points for a 10000 lb Batch', 'mill 20 mins ', 'mill 45 min', 'MILL 45 MIN', 'Mill and Deaerate for 1 hr', 'Mill for 1 hour', 'MILL FOR 1 HOUR', 'Mill for 1 Hr', 'mill for 1 HR', 'Mill for 10 mins', 'Mill for 20 mins', 'Mill for 20 mins ', 'Mill for 30 min', 'Mill for 30 mins', 'Mill for 30 Mins ', 'Mill for 30 mins ', 'Mill for 30 mins and RESAMPLE ', 'Mill for 45 mins ', 'Mill Hole Open  30mins ) ', 'Mill Longer', 'Mill Setting All the Way Open', 'MILL15 MINS ', 'Mining Grease Rework & 5%Moly LC 0 Reowrk', 'MIX & MILL 3 HRS', 'Mix & Re Circulate 1 Hr ', 'Mix & Recirculate for 15 min', 'Mix & Ricirculate 1 run through filter ', 'MIX 20 MIN', 'Mix 20mins Mill for 5 ', 'Mix 5 min Mill 5 min', 'Mix 5 min, Mill 5 min', 'Mix for 10 min', 'Mix for 30 mins ', 'Mix for 5 min and mill for 5 min', 'mix n mill for 20 mins ', 'RE SAMPLE ', 'Re Sample (Mill for 5 mins)', 'Re sample right away ', 'Re sample small can ', 'RESAMPLE', 'Resample', 'RESAMPLE ON MONDAY ', 'Sample 1 Gal can ', 'SAMPLE GALLON', 'SAMPLE GALLON FINAL TEST', 'SAMPLE MILLED & UNMILLED', 'SAMPLE UNMILLED', 'SCRAPE BLADES', 'Scrape Blades ', 'take out', 'Take Out', 'Take Out ', 'take out ', 'Takeout', 'Takeout ', 'Transfer to BK 2', 'Transfer to BK 4', 'Turn Mill On ', 'Turn Mixer Speed to Maximum Speed', 'UNMILLED SAMPLE']

REWORK = ['IXL Premium #2', 'R2 Base Grease', 'Cab O Sil TS720', 'Soft RSC Base', 'Blue Cam C 5910/680 #2', 'Thick rework (X06027G)', 'Dark White Lith', 'Red N Tacky Bin 001 Rework', 'RSC Soft Rework', 'Redtak Polyplex Haycock', 'Dark White Lith Tote 020', 'Rework Overflow BK3 with water', 'Soft Base Grease', 'Leftover in Tank', 'Tomosil PSG', 'Q Draw 7780', 'XP 900460 #1', 'Tomlith EP 2', 'Mining Grease #0', 'Rework Q11067G', 'EP Lith Rework', 'CA 207 MPC Rework', 'Complex NLGI #2', 'Rework (Tomlin 162M5C S29197G)', 'Mixed Grease drum', 'SCA 680 Rework', 'Mix Lith GR', 'Finished RSC HH #1.5 S08052H', 'MP Lith 068A #3', 'Hybrid Vault grease #0 (Z20102G)', '#1.5 rework', 'Rework 2201100042', '9810/1000 from bins ', 'B28 Base Grease (pen ~250)', 'FPL 105 #2 B07054F', 'White Lithium Rework', 'FPL 105 #2 (Bin PT03 rework)', 'Tote 004 Rework', 'Tomlith 220B #1', 'Lucas Rework', 'SMG 500 #2', 'Leftover in tank', 'Rework Tote', 'Red N Tacky Base w/water Z01001G', 'Base Grease from R2', 'Thin SMG Base', 'Tank #2 (Base w/ Moly Residue)', 'XP 9900/460 #2 (U03001G)', 'Rework Soft 2 grade (p60: 294, Y08017G)', 'Mixed Grease', 'Reladyne Mining #1 Z08074F Rweork', 'Purge Grease', 'Rework Lithium Grease', '#2.5 Rework', 'SCA 680 (LOT S04042G)', 'Soft FPL 103 #2 (X26189G pen:300)', 'XP 9900/460 #2', 'Lucas White Lithium Base (Pen ~300)', 'Rework 46-0', 'Soft SMG', 'Red N Tacky #2', '8626 V28224I', 'Rail King Re work  W18105H', 'Dark Spectrum Rework', 'Rail King Re work  T26202I', 'Red N Tacky Gritty Rework', 'Mixed Grease Drums ', 'Mixed Base Gr', 'Some kind of Lith Cmplx', 'Lucas Gritty Re Work ', 'Spectrum V  (Lot #: Z26304I)', 'Gritty/Rework', '8626 re work ', 'Red rework w/RC2540', 'B02010G', 'Liquid Re work ', 'Dust Collector', 'SCA 680 Thick A08088I', 'Rework/Gritty', 'Redtak Polyplex Overflow', 'TSG Base', 'Lucas Red N Tacky Rework A02010I', 'Termalene MP EP GR 2 2209300218', 'Burnt Base Gr from Betsy ', 'SMG 500 #2 W06021I', 'CA-207 MPC Rework', 'Red N Tacky Gritty ', 'Heavy Haul Rework Z17157I P60 :315', 'Mixed Grease ', 'FPL 105 rework ', 'K3 Grease ', 'K3 Rework', 'Unkown Bel Ray', 'Dark Rework', 'Lucas Red N Tacky Gritty', 'Lucas Red N Tacky Dark', 'All Fleet 3% #0', 'XP 9900 460 #2', '#0 and #00 Rework', '220B', '3% Moly #1 good mat', '5% 46-0 301750150526', '8626 Base Gr w/ AC 1 V28224I P60 351', '9819/3000', 'Additized Rework', 'Additized rework', 'All Fleet Mining Gfrease #1', 'All Fleet Mining Grease #1', 'Aluminium Re Work', 'ALX Rework', 'Bird B Gone REWORK (1 drum)', 'BK3 Red N Tacky #2', 'CA 207', 'CA 207 MPC ', 'Cab-O-Sil TS-720', 'Claretech Eco Curve Trackside (R17155I)', 'Claretech EcoCurve Trackside', 'Claretech EcoCurve Trackside ', 'Con Ag Re work', 'Con-Ag Moly #2 Rework', 'Cool 65 Drums Lot# U26221I', 'Cool 65 Rework', 'Cotton Picker Thick ', 'Cup Calcium Complex Rework', 'Cup grease Rework', 'Dark EP Lith REWORK', 'Dark Lithium Base Rework', 'Dark Red Bin Rework', 'Dark Termalene Rework', 'Darker Marine rework', 'EP 1 Synth GR Partial', 'EP GR LC 0 Rework', 'EP LITH rework', 'Extra LI #2 from R1 ', 'Filling Line  Rework ', 'FPL 103 Silver #0 T22197H', 'FPL 103 Silver 70 #2 ', 'FPL 103 Silver 70 #2 Rework Lot#: S13101I', 'FPL 105 #2 U23186I', 'FPL 121 Conag #1', 'FPL 121 Conag #3', 'FPL 123 #0 Rework', 'FPL 123 #0 Rework P60 #340', 'FPL 123 #1 Rework', 'FPL 123 #1 Rework  U28227I', 'FPL 123 #1.5', 'FPL 123 #2 ', 'FPL 123 CON AG WITH MOLY #2', 'FPL 123 Con-Ag with Moly', 'FPL 123 ConAg w/Moly ', 'FPL 123 Conag w/Moly #2', 'FPL 123 Fr BK1', 'FPL 800 #00 Rework', 'FPL 800 #000', 'FPL 800 #000 ', 'FPL105 black cat #2  4 Drums ', 'FPL105REWORK B23179H', 'FPL105REWORK U23186I', 'FPL105REWORK V30251I', 'HD AC 180 B Rework Flush', 'HD AC 180B #2 Rework Flush BK4', 'Heavy Haul #1.5 Rework', 'Heavy Haul Rework', 'Heavy Haul Rework Z17157I P60 :318', 'Huskey EP Lith Rework', 'Huskey Rework', 'K-3 Rework', 'K3 Re Work', 'K3 Rework ', 'Lab Additives', 'LC-0 Rework', 'Li White Rework', 'Lithium White Re work ', 'Loom Oil Base Purge T06024I', 'Lubrisilk Rework', 'Lucas Marine ', 'Lucas Marine from K2', 'Lucas Marine Rework', 'Lucas Red N Tacky #2 Dark P60: 256', 'Lucas Red N Tacky NO DYE U09082I', 'Lucas Red N Tacky R08074I', 'Lucas Red N Tacky Rework', 'Lucas Red Pre RC 2540', 'LUCAS RED REWORK', 'Lucas Red Rework', 'Lucas Red#2 T4191I', 'MIXED GREASE', 'Moly 3% Moly EP GR LC 0 pen 350 2305230179', 'Moly MP EP GR 1', 'Moly Rework', 'MOLYLUBE 3% #1', 'Molylube 3% Moly EP Gr LC 0 2305230179', 'Molylube 3% Moly EP GR Rework', 'Molylube 3%Moly EP GR LC 0', 'Molylube 3%Moly EP GR LC 0 2305230179', 'Molylube 3%Moly EP GR LC 0 P60: 355', 'Molylube 3%Moly EP GR LC 0 Rework Bin', 'Molylube 5% AC 1 (14 Fluid Bags)', 'MOLYLUBE 5% MOL;Y MOLY EP GR AC 1', 'Molylube 5%Moly EP GR AC 150-0 Rework', 'Molylube 5%Moly EP GR AC 150-0 Rework 2306210171', 'Molylube 65 Base Grease', 'Molylube MP EP GR 1 Lot 2307140088', 'Molylube Multipurp EP GR 1 Rework 2307200132', 'Partial', 'Pilot Thomas #1.5 Rework', 'Rail King  W18105H P60: 333', 'Rail King Re work ', 'Rail King Re Work ', 'Rail King Rework', 'Rail King Rework in Bins', 'Re Work', 'RE Work', 'Re WORK', 'Re Work ', 'Re work ', 'RE WORK ', 'RE woRK ', 'Re Work 2211070043', 'Re Work 2211110084', 'Re Work Batch ', 'Re work Flush ', 'Re Work fr W18105H', 'Re work from trailer', 're WORK W18105H', 'Re-Sample', 'Re-Work', 'Re-work Batch W18105H', 'Red N Tacky ', 'Red N Tacky Rework U19154I', 'Red N Tacky Rework V17137I', 'Red n Tacky Rework X15141I', 'Red N Tacky Rework X15141I', 'Redtak Polyplex #1 Rework X15140I', 'Redtak Polyplex #1 Z11063H', 'Redtak Polyplex Rework Lot: X15140I', 'RedTak Rework', 'Rework', 'REWORK', 'Rework (Molyl 5% #1.5 Flush)', 'Rework 2303020035', 'Rework 2303100093', 'Rework 2303200169', 'rework 2306010004', 'Rework Base A07051H', 'Rework fr W18105H', 'Rework Grease', 'Rework oil bleed', 'Rework Poly Urea', 'RP Thermasil Lot# 2303170156', 'RP Thermasil T-100 Rework', 'RP Ultra Performance GR Rework Lot 2308230196', 'RP UPG Rework Lot 2308230196', 'RSC HH P60 288', 'RSC RAIL KING HL REWORK', 'Silica Rework', 'Silver 70 #1', 'Silver 70 Rework P60: 378', 'Silver Bin Rework', 'Silver re work  ', 'SMG 500 #2 Lot Z31258I', 'SMG 500 #2 Rework ', 'SMG 500 with Char', 'SMG 500B#2 Rework', 'SMG RE WORK ', 'Soft Loom Oil Base Grease', 'Tan Al Rework', 'Tan Lith Rework', 'Termalene #1 Rework', 'Termalene MP EP GR 1 Rework', 'Termalene MP EP GR 1 Rework w/o VL 622', 'Termalene MP EP GR 2 Lot 2308040052', 'Termalene MP EP GR 2 Rework', 'Termalene Re Work ', 'Termalene Rework', 'Termelene MP EP GR 2 Rework 2209300218', 'Termelene MP EP GR 2 Rework 2308030030', 'Thermasil Re work ', 'Thermasil T100', 'Thick FPL 105 fr last batch ', 'Thin Loom', 'Thin Loom Base Rework U16128I', 'Thin Rework', 'Thin SMG Base W25213I', 'Tomlin Syn 460/3 Partial V03019N ', 'Tomlith 220B ', 'Torco Rework ', 'Trackside Rework', 'Tuff Coat 20', 'Tuff Coat 20 ', 'Tuff Coat Flush w/ Silica', 'White Li Re-Work', 'White Re Work ', 'White Rework', 'XE 9800 #1', 'XP 9900/460 # 2', 'Pro One ESP ']

ELCOS476229 = ['Elco 47629 ', 'Elco47629', 'Elcon 47629','ELCO 47629', 'Elco 47629']

LIQUID_ADDITIVES = ['Functional V188', 'Ethanol Solution 3', 'Hitex 4313', 'Cuvan 829', 'Functional V-425', 'Additin RC 2540 / TPS 44', 'Additin 3775', 'Irgabox L150', 'PB 1300', 'RC 9300', 'OMS ', 'OMS Spirits Oderless', 'Nanoflon M020', 'XL10', 'OMS', 'Ocopol AZ/ Vanlube AZ', 'Engine Life Treatment', 'Tergitol 15-S-3 ', 'Citronella Essential Oil', 'Functional V 176', '"Engine Oil Treatment"', 'Songnox 1035 (Vanlube AZ)', 'Maxi Lube Additive', 'DSS (Disodium Sebacate)', 'Actopol AZ / Vanlube AZ', 'Vanlube 81', 'Tinuvin 329', 'Emersol 213', 'Green Dye', 'Octopol SIB', 'Octolube 229EP', 'Functional V-188', 'Doverlube SP44', 'Na Lube A0242 (if no Songnox 1035)', 'Liquid Oil Blue', 'Peppermint Essential Oil', 'NA SUL CA 1259', 'Functional V425', 'Amorcal 400M', 'Chimassorb 81', 'Gama-Sperse 80', 'Infineum C9340', 'KCORR G-1340', 'LZ ADX511C', 'Markphos DPDP', 'Mayco Base 1210', 'SAG 100','Additin RC 2540', 'Additin RC 2540 \\', 'Additin RC 2540/TPS 44', 'Additin RC 2540/TPS44', 'Additin RC 3045', 'Additin RC-2540', 'Additin RC-2540/ TPS 44 ', 'Additin RC2540', 'Anti Foam ', 'anti foam ', 'Bonderite S-AD 1275', 'Bonderite SAD 1275', 'Calcinate 400CLR', 'Calcinate C 400', 'Calcinate C-400', 'Calcinate C400', 'CaLCINATE c400', 'Calcinate C400 ', 'Calcinate Counter Rust CA400', 'Carbowax MPEG 550', 'Chemtura-CLR 400', 'Counter Rust ', 'Counter Rust 7155', 'Counter Rust CA-42', 'Counter Rust CA-43 / CS-4', 'Counter Rust CA42', 'Cuvan', 'Cuvan ', 'Cuvan  826', 'Cuvan 826', 'Cuvan 826 ', 'Cuvan 828', 'Dehylube 4030', 'Desilube 88', 'Doverlube SP 44', 'Elco 103', 'Elco 105', 'Elco 105 ', 'Elco 223', 'Elco 233', 'Elco105', 'EM500', 'Emerox 1110', 'Emersol 875', 'Emery 658', 'Functional V-174', 'Functional V-176', 'Functional V-178', 'Functional V-184', 'Functional V176', 'Hi Tec 4313', 'Hitec 4313', 'Hitec 611', 'Htec 4313', 'Igrament 39', 'Irgamet 39', 'Irganox 39 ', 'Irganox L115', 'Irganox L150', 'Irganox L150 ', 'Irganox L57', 'Irganox L57 ', 'Irgonox L57', 'Lubrizol 5034A', 'Lubrizol 5080A', 'Lubrizol 5346', 'Lubrizol 889D', 'Lubrizol ADX 511C', 'Lumisorb SMO K', 'Mobil AN-917', 'Mobil SS-2300', 'NA Sul 729', 'Na-Lube KR 019', 'NA-Sul 729', 'NA-SUL CA-1259', 'Octopo AZ/Vanlube AZ ', 'Octopol AD / Vanlube 73', 'Octopol AD/ Vanlube 73', 'Octopol AD/Vanlube 73', 'Octopol AZ', 'Octopol AZ / Vanlube AD', 'Octopol AZ / Vanlube AZ', 'Octopol AZ/ Vanlube AZ', 'Octopol AZ/ vanlube AZ', 'Octopol AZ/ Vanlube AZ 73', 'Octopol AZ/Vanlube AZ', 'Octopol AZ/Vanlube AZ ', 'Octopol Bis DMTD', 'Octopol DMTD', 'Octopol LTX 001501', 'Octopol PTB', 'Paratac', 'Paratac ', 'Paratac\\', 'Paratc ', 'ParOil 45', 'Priolube 2500', 'Q DRAW 7780', 'Q-Draw 7780', 'RC 2540', 'RC 2540 / TPS 44', 'rc2540', 'RC2540', 'Shamrock Nanoflon MO20', 'Sulfomed A ', 'Sulfomed A-500', 'Sulfomed A-500 / EZ Mulz 50', 'Sulfomed A-500 / EZ Mulz 500', 'Sulfomed A-500 / EZ-Mulz 500', 'Sulfomed A500', 'Thermax Grease Add Base A', 'Thermax Grease Add Base A ', 'Thermax Grease Add Base B', 'Tomlin Anti Foam', 'Tomlin Antifoam', 'Tomlin AntiFoam', 'Tomlin XL 10 ', 'Tomlin XL-10', 'Tomlin XL10', 'TPS', 'TPS 44', 'tps 44', 'TPS 44 ', 'TPS 44 / RC 2540', 'TPS44', 'tps44', 'UPG Add Base', 'UPG Add Base ', 'Vanlube', 'Vanlube ', 'Vanlube 622', 'Vanlube 73', 'Vanlube 7723', 'Vanlube 9123', 'Vanlube AZ', 'vaNLUBE az', 'Vanlube RI -G', 'Vanlube RI-G', 'Vanlube SB', 'Vanlube TK-100', 'Kendex 0834 (color)', 'Unisol Liquid Blue', 'Unisol Liquid Blue ', 'Unisol Liquid Green', 'Unisol Liquid Green ', 'Unisol Liquid Purple ', 'Unisol Liquid Purple BRHF', 'Unisol Liquid Red', 'Unisol Liquid Red ']
SOLID_ADDITIVES = ['Tech Grade Moly', 'Molybdenum Disulfide', 'Totanium Dioxide', 'Tech Fine Moly', 'Ciba L160', 'Titanium White', 'Beta Carotene', 'Charmax Antimony Trioxide', 'Asbury 4420 graphite', 'Molysulfide Tech Fine Mix', 'Octopol Bis DMTD / Vanlube', 'Xtendra (BHT)', 'VutaCal PCC', 'Carbon BlaCK', 'AB 100%-01', 'Beta Carotene ', 'Asbury Graphite 4420', 'acetylene Black 100%', 'Molysulfide Tech Fine', 'Asbury 4420 Graphite', 'Cuba L160', 'Titanium Dioxide R996', 'Glitter', 'Zinc Oxide USP 104', 'Thixocal 300A', 'AB 110% or Ace Carbon Black ', 'Tech FINE Moly', 'AB 75 % ', 'Songnox 1010', 'Titanium Dioxide ', 'Asbury Synthetic Graphite 1176', 'Asbury 4680', 'Vulcan XC 72R', 'Keyplast Yellow Fluorescent', 'Capsaicin 95% Powder', 'Molysulfide Tech Fine Grade ', 'Copper Powder MD 3700', 'PCC', 'Capsaicin 95% powder', 'Shamrock Fluoro E', 'Vicron 1515 ', 'Extra Fine Silver (Al paste)', 'Vulcan XC-72R', 'Vulcan XC-72R', 'AB 75%-01', 'Acetylene Balck 50%', 'Acetylene Black', 'Acetylene Black ', 'Acetylene Black 50%', 'Acetylene Black 50%-01', 'Acetylrne Black  50%', 'Aluminum Paste ', 'Asbury 1461 Graphite', 'Asbury 4680 graphite', 'Asbury 4680 Graphite', 'Asbury 4680 Graphite ', 'Asbury Synthetic Graphite 1', 'Ashbury 4420', 'Ashbury Graphite 4680', 'Beta Carotene 8030', 'BHT', 'Bismuth Trioxide (Bi2O3) Te', 'Carbon Black Conductex SC', 'Charmax SB', 'Charmax SB ', 'Curbotec 7005 Copper', 'Fluoro 70', 'Fluoro E', 'Fuoro E', 'Graphite 4420', 'Graphite 4680', 'Graphite 5039', 'Hystrene 3002', 'Moly Tech Fine G', 'Moly Tech Grade', 'Molysulfide Tech Fine G', 'Molysulfide Tech Fine Grade', 'Molysulfide Tech Fine Grade G', 'Molysulfide Tech Fine Moly', 'Molysulfide Tech grade', 'Molysulfide Tech Grade', 'Molysulfide tech grade', 'Molysulfide Tech Grade ', 'Molysulfide Tech Grade Fine G', 'Molysulfide Techincal Fine G', 'Molysulfide Techinical Fine G', 'Molysulfide Technicak Fine G', 'Molysulfide Technical Fine G', 'Molysulfide Technical Fine G ', 'Molysulfide Technical Fine Grade', 'Molysulfide Technical Fine Grade ', 'Molysulfide Technical Grade', 'SDF 5-203 Aluminum Paste', 'SDF 5-901 Aluminum Paste', 'Songnox 1035', 'Songonx 1035', 'Tech Fine Moly ', 'Titanium Dioxide', 'Titanium Dioxide R-996', 'Titanium Dioxide R-996/ Kro', 'Titanium Dioxide TR 93', 'Titanium Dioxide TR-93', 'Titanium Dioxide TR-996', 'Titantiu Dioxide TR-93', 'Titianium Dioxide TR-93', 'Vanlube 829', 'VanLube 829', 'Vanlube 829 ', 'ViaCal PCC', 'Vicality PCC', 'Vicrn 1515', 'Vicron', 'Vicron 15-15', 'Vicron 15-15 HG', 'Vicron 1515', 'VICRON 1515', 'Vicron1515', 'VitaCal / ViCality SG', 'Vitacal PCC', 'VitaCal PCC', 'VitaCal PCC / Vicality SG', 'VitaCal PCC / ViCality SG', 'Vitacal PCC / Vitacality SG', 'VitaCal PCC / Vitacality SG', 'Vitacal PCC / Vitaclity SG', 'VitaCal PCC/ViCality SG', 'VitalCal PCC', 'Vitalcal PCC ', 'VitalCal PCC ', 'Vitalcal PCC / Vicality SG ', 'Vitality PCC', 'Vucron 1515', 'Zinc Dioxide', 'Zinc Oxide', 'ZINC OXIDE', 'Zinc Oxide ']

REACTANTS = ['Boric Acid', 'Boric Acid (TP)', 'Caprylic V', 'Cenwax A', 'Cenwax A (12 HSA)', 'Cenwax A (12HSA)', 'Cenwax A or 12 HAS', 'Cenwax G', 'DDBSA', 'Hexylene Glycol', 'Lithium Hydroxide', 'Lithium Hydroxide Monohydrate', 'MicroCal HF', 'Microthene F', 'Microthene F FN 510-00', 'MV-200 Lime', 'Phosphoric Acid-75%', 'Propylene Glycol', 'Unvar LAS 98 L2P', 'Acetic Acid', 'Boric Acid', 'Boric Acid (TP)', 'Caprylic V', 'Cenwax A', 'Cenwax A (12 HSA)', 'Cenwax A (12HSA)', 'Cenwax A or 12 HAS', 'Cenwax G', 'DDBSA', 'Hexylene Glycol', 'Lithium Hydroxide', 'Lithium Hydroxide Monohydrate', 'MicroCal HF', 'Microthene F', 'Microthene F FN 510-00', 'MV-200 Lime', 'Phosphoric Acid-75%', 'Propylene Glycol', 'Unvar LAS 98 L2P']
TRASH = ['Baragel +water changes', 'Bentone 34 GG', '12-Hydroxy Stearic Acid', '12 HAS', 'Microcal HF', 'LZ 2002', 'Acetic Acid Glacial 99%', 'Hydrogenated Castor Oil', 'Emersol 3875', 'null', '', '', 'water (tap)', 'Water H2O', 'Baragel 3000', 'Bentone 34', 'LPA 170', '0', '1', '2', '3', '4', '5', '290', '2539', '2751', '2871', '2951', '2204280211', '2211080072', '2211286176', '???', 'Final Ponds', 'Initial Amount', 'Other Assumption:', 'p60 =', 'p60:', 'p60: 271 @78', 'p60: 300-330', 'Pounds Needed', 'POUNDS OF OIL TO ADD', 'Subtotal', 'Test 12', 'Test 13', 'Test 14', 'Test 15', 'Test 16', 'Test 17', 'Test 18', 'Test 19', 'Test 20', 'Total', 'unworked1', 'Additives', 'ADDITIVES ', 'Components', 'PX 3841', 'Tank Base Pan', 'Water', 'water']


def run_new_read():
	##############
	#@ no paramaters everything implied
	# @ returns:
	#   
	

	companys, names, reactors, lots, date_apr, base_gr, weld_load, base_pen = [], [], [], [], [], [], [], []
	base_grease_pct, base_oil_pct = [], []
	silica_pct = []
	rework_pct = []
	rework_factors = []
	sum_of_percents = []

	solid_additive_pct, liquid_additive_pct = [], []
	elco_complex = []
	flags_for_removal = []
	all_other_additives = set()
	count = 0

	companies = os.listdir('QC Analysis')
	companies.remove('.DS_Store')
	#print(companies, '\n'+str(len(companies)))

	for company in companies:
		workbooks = os.listdir(f'QC Analysis/{company}')
		#print(company, workbooks)

		for workbook in workbooks:
			print(workbook)
			if 'DS_Store' not in workbook:
				wb = load_workbook(f'QC Analysis/{company}/{workbook}', data_only=True)
				worksheets = wb.sheetnames
				print(workbook, worksheets)

				for worksheet in worksheets:
					if ("Outline" and "Overview" and "Notes" and "Sheet") not in worksheet:
						print(worksheet)
						# load worksheet
						ws = wb[worksheet]
						# reset flags
						elco_flag = False
						flag_for_removal = False
						# take easy reads
						names.append(ws['B1'].value)
						reactors.append(ws['B5'].value)
						lots.append(ws['B2'].value)
						date_apr.append(ws['B4'].value)
						weld_load.append(ws['B6'].value)
						companys.append(ws['B3'].value)
						base_pen.append(ws['U6'].value)

						#reset things for each sheet (batch)
						bgp = 0  # Base Grease Percentage
						bop = 0  # Base Oil Percentage
						lap = 0  # Liquid Additives Percentage
						sap = 0  # Solid Addditives Percentage
						fop = 0  # Flush Oil Percentage
						sip = 0  # Si Percentage
						tbp = 0  # Thicc Bois Percentage
						trp = 0  # Total Rework Percentage
						prp = 0  # unsure but unused maybe i'll remember someday
						empty_count = 0
						base_greece = ""
						rework_adjustment_factor = 1
						
						# account for differences in formatting over time
						if str(ws[f'B11'].value) == 'Base Material':
							interval = [32, 33] + [x for x in range(42, 54)]
						else:
							interval = [x for x in range(11, 31)]

						# loop over components list
						for i in interval:

							# make sure we dont loop too much
							if str(ws[f'B{i}'].value) == ("" or "0") :
								empty_count += 1
							if empty_count >= 2:
									break
							
							# check if cell is base grease and add it to base pct
							if str(ws[f'B{i}'].value).upper() in BASE_GREASES or str(ws[f'B{i}'].value) in ALT_BASED_GR:
								try:
									bgp += float(ws[f'O{i}'].value)
									base_greece += str(ws[f'B{i}'].value)
								except(TypeError, ValueError):
									bgp += 0
									base_greece += str(ws[f'B{i}'].value)
									flag_for_removal = True
						
							# check for base oils
							elif str(ws[f'B{i}'].value).upper() in BASE_OILS or str(ws[f'B{i}'].value) in ALT_BASED_OIL:
								try:
									bop += float((ws[f'O{i}'].value))
								except(TypeError, ValueError):
									bop += 0
									flag_for_removal = True
							# elco complex check
							elif str(ws[f'B{i}'].value) in ELCOS476229:
								elco_flag = True
								try: 
									lap += (float(ws[f'O{i}'].value))
								except(ValueError, TypeError):
									lap += 0
									flag_for_removal = True
							
							# liquid additives pipeline
							elif str(ws[f'B{i}'].value) in LIQUID_ADDITIVES:
								try:
									lap += float(ws[f'O{i}'].value)
								except(TypeError, ValueError):
											lap += 0
											flag_for_removal = True
							
							# solid additives filter
							elif str(ws[f'B{i}'].value) in SOLID_ADDITIVES:
								try:
									sap += float(ws[f'O{i}'].value)
								except(TypeError, ValueError):
											sap += 0
											flag_for_removal = True
							
							# flush drainage
							elif str(ws[f'B{i}'].value) in FLUSH:
								try:
									fop += float(ws[f'O{i}'].value)
								except(TypeError, ValueError):
											fop += 0
											flag_for_removal = True
								bop += fop * 0.9
								bgp += fop * 0.1
							
							# silica processor
							elif str(ws[f'B{i}'].value) in SILICA:
								try:
									sip += float(ws[f'O{i}'].value)
								except(TypeError, ValueError):
											sip += 0
											flag_for_removal = True
							
							# polymer and resin squeeze
							elif str(ws[f'B{i}'].value) in THICC_BOIS:
								try:
									tbp += float(ws[f'O{i}'].value)
								except(TypeError, ValueError):
											tbp += 0
								bop += tbp * 0.9
								rework_adjustment_factor = rework_adjustment_factor * (1 - (tbp/100 - tbp/100 * 0.9))
							
							# rework 
							elif str(ws[f'B{i}'].value) in REWORK:
								try:
									trp += float(ws[f'O{i}'].value)
								except(TypeError, ValueError):
											trp += 0
											flag_for_removal = True
								rework_adjustment_factor = rework_adjustment_factor * (1 - trp/100)

							# flag dogshit for removal after next iteration
							elif (str(ws[f'B{i}'].value) or str(ws[f'J{i+31}'].value)) in INSTRUCTIONS:
								pass
							elif (str(ws[f'B{i}'].value) or str(ws[f'J{i+31}'].value)) in REACTANTS:
								flag_for_removal = True
							elif (str(ws[f'B{i}'].value) or str(ws[f'J{i+31}'].value)) in TRASH:
								flag_for_removal = True
							#document the ones that got away
							else:
								all_other_additives.add(ws[f'B{i}'].value)


						try: 
							base_grease_pct.append(bgp/rework_adjustment_factor)
							base_oil_pct.append(bop/rework_adjustment_factor)
							liquid_additive_pct.append(lap/rework_adjustment_factor)
							solid_additive_pct.append(sap/rework_adjustment_factor)
							silica_pct.append(sip/rework_adjustment_factor)
							sum_of_percents.append((bgp + bop + lap + sap + sip) / rework_adjustment_factor)
						except(ZeroDivisionError):
							flag_for_removal = True
							base_grease_pct.append(0.0)
							base_oil_pct.append(0.0)
							liquid_additive_pct.append(0.0)
							solid_additive_pct.append(0.0)
							silica_pct.append(0.0)
							sum_of_percents.append(0.0)
						base_gr.append(base_greece)
						elco_complex.append(elco_flag)
						flags_for_removal.append(flag_for_removal)
						rework_pct.append(trp)
						rework_factors.append(rework_adjustment_factor)
						count += 1
						print(count)

	grease_array = [names, date_apr, lots, reactors, companys, weld_load, base_oil_pct, base_grease_pct, liquid_additive_pct, solid_additive_pct, silica_pct, sum_of_percents, elco_complex, flags_for_removal, base_gr, rework_pct, rework_factors]
					
	return grease_array, list(all_other_additives)

def make_new_master(X, iteration):
	new_wb = Workbook()
	master_ws = new_wb.active
	master_ws['B1'] = "Name" #0
	master_ws['C1'] = "Date Approved" #1
	master_ws['D1'] = "Lot Number" #2
	master_ws['E1'] = "Kettle/Finisher" #3
	master_ws['F1'] = "Company" #4
	master_ws['G1'] = "Weld Load" #5
	master_ws['H1'] = "Base Oil %" #6
	master_ws['I1'] = "Base Grease %" #7
	master_ws['J1'] = "Liquid Additives %" #8
	master_ws['K1'] = "Solid Additives %" #9
	master_ws['L1'] = "Silica %" #10
	master_ws['M1'] = "Total %" #11
	master_ws['N1'] = "Elco Complex" #12
	master_ws['O1'] = "Flag for Removal" #13
	master_ws['P1'] = "Base Grease" #14
	master_ws['Q1'] = "Rework"
	master_ws['R1'] = "factor"

	for i in range(len(X[0])):
		master_ws[f'B{i+2}'] = X[0][i]
		master_ws[f'C{i+2}'] = X[1][i]
		master_ws[f'D{i+2}'] = X[2][i]
		master_ws[f'E{i+2}'] = X[3][i]
		master_ws[f'F{i+2}'] = X[4][i]
		master_ws[f'G{i+2}'] = X[5][i]
		master_ws[f'H{i+2}'] = X[6][i]
		master_ws[f'I{i+2}'] = X[7][i]
		master_ws[f'J{i+2}'] = X[8][i]
		master_ws[f'K{i+2}'] = X[9][i]
		master_ws[f'L{i+2}'] = X[10][i]
		master_ws[f'M{i+2}'] = X[11][i]
		master_ws[f'N{i+2}'] = X[12][i]
		master_ws[f'O{i+2}'] = X[13][i]
		master_ws[f'P{i+2}'] = X[14][i]
		master_ws[f'Q{i+2}'] = X[15][i]
		master_ws[f'R{i+2}'] = X[16][i]

	new_wb.save(f"master_matrix_{iteration}.xlsx")
	
def get_pens():
	companies = os.listdir('QCTestResults')
	companies.remove('.DS_Store')

	pen_list = [[],[]]
	count = 0
	for company in companies:
		workbooks = os.listdir(f'QCTestResults/{company}')
		for workbook in workbooks:
			if 'DS_Store' not in workbook:
					wb = load_workbook(f'QCTestResults/{company}/{workbook}', data_only=True)
					worksheets = wb.sheetnames
					for worksheet in worksheets:
						ws = wb[worksheet]
						#print(worksheet)
						max_row = ws.max_row
						for i in range(1, max_row+1):
							pen_list[0].append(ws[f'B{i}'].value)
							pen_list[1].append(ws[f'D{i}'].value)
							count += 1
							#print(count)
	return pen_list


grease_X, others = run_new_read()
make_new_master(grease_X, 'final')
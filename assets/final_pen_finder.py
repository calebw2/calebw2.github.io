from openpyxl import Workbook, load_workbook
import os

def get_pens():
	companies = os.listdir('QCTestResults')
	companies.remove('.DS_Store')

	pen_list = [[],[]]
	count = 0
	for company in companies:
		workbooks = os.listdir(f'QCTestResults/{company}')
		for workbook in workbooks:
			if 'DS_Store' not in workbook:
					wb = load_workbook(f'QCTestResults/{company}/{workbook}', data_only=True)
					worksheets = wb.sheetnames
					for worksheet in worksheets:
						ws = wb[worksheet]
						print(worksheet)
						max_row = ws.max_row
						for i in range(1, max_row+1):
							pen_list[0].append(ws[f'B{i}'].value)
							pen_list[1].append(ws[f'D{i}'].value)
							count += 1
							print(count)
	return pen_list


print(len(pen_list[0]), len(pen_list[1]))

print(pen_list)